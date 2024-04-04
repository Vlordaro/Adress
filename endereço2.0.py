import openpyxl

def formatar_endereco(endereco, numero):
    endereco_formatado = endereco.upper() + " Nº" + str(numero)
    return endereco_formatado

# Abrir o arquivo Excel
caminho_arquivo = r'C:\Users\Vinicius L\Desktop\Controle BK_2024 (2.0).xlsx'
nome_planilha = 'Lojas'  
wb = openpyxl.load_workbook(caminho_arquivo)
sheet = wb[nome_planilha]

# Definir uma variável para controlar se o usuário deseja preencher outro endereço
continuar_preenchendo = True

# Loop para continuar perguntando ao usuário se ele deseja preencher outro endereço
while continuar_preenchendo:
    # Ler o número buscado na coluna "BKN"
    numero_procurado = input("Digite o número buscado na coluna BKN: ")
    
    # Verificar se o número existe na planilha
    numero_encontrado = False
    for linha in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for celula in linha:
            valor_bkn = str(celula.value)  # Convertendo para string
            if valor_bkn == numero_procurado:  
                endereco = input("Digite o endereço: ")
                numero = input("Digite o número do endereço: ")

                # Formatar o endereço
                endereco_formatado = formatar_endereco(endereco, numero)

                # Preencher a coluna "Lojas" na mesma linha com o endereço formatado
                linha_desejada = celula.row
                sheet.cell(row=linha_desejada, column=3, value=endereco_formatado)  

                # Salvar as alterações no arquivo Excel
                wb.save(caminho_arquivo)
                print("Endereço formatado foi inserido na linha", linha_desejada)
                numero_encontrado = True
                break
        if numero_encontrado:
            break
    else:
        print("Número não encontrado na coluna BKN.")

    # Perguntar ao usuário se ele deseja preencher outro endereço
    resposta = input("Deseja preencher outro endereço? (S/N): ")
    if resposta.upper() != 'S':
        continuar_preenchendo = False
